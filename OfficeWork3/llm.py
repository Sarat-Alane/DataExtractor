# llm.py

import os
import pandas as pd
from dotenv import load_dotenv
import google.generativeai as genai
from openpyxl import load_workbook, Workbook


# ---------------------------------------------------------
# Load API Key
# ---------------------------------------------------------
load_dotenv()
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

model = genai.GenerativeModel("models/gemini-2.5-flash")


# ---------------------------------------------------------
# Read and preprocess text
# ---------------------------------------------------------
def load_extracted_text(file_path: str) -> str:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"❌ File not found: {file_path}")

    with open(file_path, "r", encoding="utf-8") as f:
        raw_text = f.read()

    cleaned = raw_text.replace("=" * 70, "").replace("-" * 50, "")
    cleaned = "\n".join([line.strip() for line in cleaned.split("\n") if line.strip()])

    return cleaned


# ---------------------------------------------------------
# ❶ Extract ONLY table as CSV (your original function)
# ---------------------------------------------------------
def extract_table_llm(text: str) -> str:
    prompt = f"""
You are a table extraction engine.

Your task:
Extract *only the actual table data* from the OCR text below.

Rules:
1. Do NOT include totals, headers, bank details, addresses, or summaries.
2. Detect all tables present. Return them one by one.
3. Format output as pure CSV only. No markdown. No commentary.
4. Cells in a row must be separated with commas.
5. Rows separated by newlines.
6. If multiple tables, separate with a blank line.
7. Do not invent missing values. Use only text found in OCR.

OCR TEXT:
{text}
"""

    response = model.generate_content(prompt)
    return response.text.strip()


# ---------------------------------------------------------
# ❷ Extract Trader Name, Invoice Date, Billing Period (same prompt)
# ---------------------------------------------------------
def extract_metadata_llm(text: str):
    prompt = f"""
Extract the following metadata from the OCR text:

1. Trader Name (supplier/vendor name)
2. Invoice Date (must be a SINGLE date, not a range)
3. Billing Period (must be a DATE RANGE)

Return output in this EXACT format:

Trader Name: <value>
Invoice Date: <value>
Billing Period: <value>

Do NOT add anything else.

OCR TEXT:
{text}
"""

    response = model.generate_content(prompt)
    output = response.text.strip()

    trader, invoice_date, billing_period = "", "", ""

    for line in output.split("\n"):
        line = line.strip()
        if line.startswith("Trader Name:"):
            trader = line.replace("Trader Name:", "").strip()
        elif line.startswith("Invoice Date:"):
            invoice_date = line.replace("Invoice Date:", "").strip()
        elif line.startswith("Billing Period:"):
            billing_period = line.replace("Billing Period:", "").strip()

    return trader, invoice_date, billing_period


# ---------------------------------------------------------
# Build DataFrame from CSV + add metadata columns
# ---------------------------------------------------------
def build_dataframe_from_csv(csv_text: str, trader: str, invoice_date: str, billing_period: str) -> pd.DataFrame:
    # Multiple tables separated by blank line
    tables = csv_text.strip().split("\n\n")
    dfs = []

    for tbl in tables:
        if not tbl.strip():
            continue
        lines = tbl.strip().split("\n")
        rows = [line.split(",") for line in lines]
        df = pd.DataFrame(rows)

        # Add metadata columns (repeated for each row)
        df["Trader Name"] = trader
        df["Invoice Date"] = invoice_date
        df["Billing Period"] = billing_period

        dfs.append(df)

    if not dfs:
        return pd.DataFrame()

    return pd.concat(dfs, ignore_index=True)


# ---------------------------------------------------------
# Append DataFrame to a single Excel file (one sheet)
# ---------------------------------------------------------
def append_to_excel(df_new: pd.DataFrame, output_path: str = "table_output.xlsx", sheet_name: str = "All_Invoices"):
    if df_new.empty:
        print("⚠️ No rows to append (empty DataFrame from CSV). Skipping Excel write.")
        return

    if not os.path.exists(output_path):
        # Create new workbook and write from scratch
        df_new.to_excel(output_path, sheet_name=sheet_name, index=False, header=False)
        print(f"📄 Created new Excel file: {output_path}")
        return

    # Append with openpyxl (no need to read into pandas again)
    wb = load_workbook(output_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Append each row of df_new
    for row in df_new.itertuples(index=False, name=None):
        ws.append(list(row))

    wb.save(output_path)
    print(f"📄 Appended {len(df_new)} rows to: {output_path} (sheet: {sheet_name})")


# ---------------------------------------------------------
# Main execution (used per-image in ocr.py)
# ---------------------------------------------------------
def run_llm_table_extraction(
    input_file: str = "extracted_text.txt",
    output_xlsx: str = "table_output.xlsx",
):
    print("📄 Loading extracted text...")
    text = load_extracted_text(input_file)

    print("🤖 Extracting TABLE via Gemini...")
    csv_output = extract_table_llm(text)

    print("🤖 Extracting METADATA via Gemini...")
    trader, invoice_date, billing_period = extract_metadata_llm(text)

    print("\n=== METADATA EXTRACTED ===")
    print("Trader Name:", trader)
    print("Invoice Date:", invoice_date)
    print("Billing Period:", billing_period)

    print("\n📝 Building DataFrame with metadata columns...")
    df_final = build_dataframe_from_csv(csv_output, trader, invoice_date, billing_period)

    print("📦 Appending table to cumulative Excel file...")
    append_to_excel(df_final, output_xlsx)

    print("\n🎉 Table extraction + metadata append complete for this OCR text!")


if __name__ == "__main__":
    # This still supports the old "single text file" flow if you want to run directly.
    run_llm_table_extraction()
